#!/usr/bin/env python3
"""
Dispatch Tracking Frame for Cylinder Management System
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
from database import dispatch_cylinders, return_cylinders, get_all_dispatches, get_all_customers, get_cylinders_by_status, get_dispatched_cylinders_by_dc, generate_dc_number, get_connection
from models.dispatch import Dispatch
from models.customer import Customer
try:
    from openpyxl import Workbook
except ImportError:
    messagebox.showerror("Missing Library", "openpyxl is required for Excel export. Please install it with: pip install openpyxl")
    Workbook = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
except ImportError:
    messagebox.showerror("Missing Library", "reportlab is required for PDF generation. Please install it with: pip install reportlab")
    SimpleDocTemplate = None

class DispatchTrackingFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.dispatches = []
        self.customers = []
        self.available_cylinders = []
        self.selected_items = set()  # For checkbox selection
        self.cyl_history_selected = set()  # For available cylinders history selection
        self.load_customers()
        self.load_available_cylinders()
        self.create_widgets()
        self.load_dispatches()
        self.load_available_cylinders_history()

    def generate_dc_number(self):
        """Generate a unique DC number."""
        return generate_dc_number()

    def create_widgets(self):
        """Create dispatch tracking widgets."""

        # Main container with tabs
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        # Tab 1: Operations (Dispatch/Return)
        operations_tab = ttk.Frame(self.notebook)
        self.notebook.add(operations_tab, text="Operations")

        # Left panel - Dispatch/Return operations with scroll
        left_panel = ttk.LabelFrame(operations_tab, text="Operations")
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 2), pady=5)
        
        # Canvas and scrollbar for scrollable operations section
        canvas = tk.Canvas(left_panel)
        scrollbar = ttk.Scrollbar(left_panel, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Dispatch section
        dispatch_frame = ttk.LabelFrame(scrollable_frame, text="Dispatch Cylinders", padding=5)
        dispatch_frame.pack(fill=tk.X, padx=5, pady=5)

        # Grid layout for dispatch form
        tk.Label(dispatch_frame, text="Customer:").grid(row=0, column=0, padx=5, pady=3, sticky="w")
        self.customer_var = tk.StringVar()
        self.customer_combo = ttk.Combobox(dispatch_frame, textvariable=self.customer_var,
                                           values=[f"{c[0]} - {c[1]}" for c in self.customers],
                                           state="readonly", width=28)
        self.customer_combo.grid(row=0, column=1, padx=5, pady=3, sticky="ew")

        tk.Label(dispatch_frame, text="Dispatch Date:").grid(row=1, column=0, padx=5, pady=3, sticky="w")
        self.dispatch_date_entry = tk.Entry(dispatch_frame, width=30)
        self.dispatch_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.dispatch_date_entry.grid(row=1, column=1, padx=5, pady=3, sticky="ew")

        tk.Label(dispatch_frame, text="DC Number:").grid(row=2, column=0, padx=5, pady=3, sticky="w")
        self.dc_number_var = tk.StringVar()
        self.dc_number_entry = tk.Entry(dispatch_frame, textvariable=self.dc_number_var, width=30)
        self.dc_number_entry.grid(row=2, column=1, padx=5, pady=3, sticky="ew")
        # Auto-fill with next available DC number
        self.dc_number_entry.insert(0, self.generate_dc_number())

        tk.Label(dispatch_frame, text="Grade:*").grid(row=3, column=0, padx=5, pady=3, sticky="w")
        self.grade_entry = tk.Entry(dispatch_frame, width=30)
        self.grade_entry.grid(row=3, column=1, padx=5, pady=3, sticky="ew")

        tk.Label(dispatch_frame, text="Available Cylinders:").grid(row=4, column=0, padx=5, pady=3, sticky="nw")
        # Frame for listbox and scrollbar
        cylinder_frame = tk.Frame(dispatch_frame)
        cylinder_frame.grid(row=4, column=1, padx=5, pady=3, sticky="ew")
        # Listbox for multiple cylinder selection
        self.cylinder_listbox = tk.Listbox(cylinder_frame, selectmode=tk.MULTIPLE, height=4, width=32)
        self.cylinder_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cylinder_scrollbar = tk.Scrollbar(cylinder_frame, orient=tk.VERTICAL, command=self.cylinder_listbox.yview)
        self.cylinder_listbox.config(yscrollcommand=cylinder_scrollbar.set)
        cylinder_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.cylinder_listbox.bind('<<ListboxSelect>>', lambda e: self.on_available_cylinder_select())
        self.cylinder_listbox.bind('<<ListboxUnselect>>', lambda e: self.on_available_cylinder_deselect())
        # Populate listbox
        for cylinder in self.available_cylinders:
            self.cylinder_listbox.insert(tk.END, f"{cylinder[0]} - {cylinder[1]} ({cylinder[2]})")

        tk.Label(dispatch_frame, text="Manual Cylinder IDs:").grid(row=5, column=0, padx=5, pady=3, sticky="w")
        self.manual_cylinder_entry = tk.Entry(dispatch_frame, width=30)
        self.manual_cylinder_entry.grid(row=5, column=1, padx=5, pady=3, sticky="ew")
        self.manual_cylinder_entry.bind('<Return>', lambda e: self.update_selected_cylinders())
        tk.Label(dispatch_frame, text="(Comma-separated IDs)", font=("Arial", 8)).grid(row=6, column=1, padx=5, pady=0, sticky="w")

        tk.Label(dispatch_frame, text="Selected Cylinders:").grid(row=7, column=0, padx=5, pady=3, sticky="nw")
        # Frame for listbox and scrollbar
        selected_frame = tk.Frame(dispatch_frame)
        selected_frame.grid(row=7, column=1, padx=5, pady=3, sticky="ew")
        self.selected_cylinders_listbox = tk.Listbox(selected_frame, height=3, width=32)
        self.selected_cylinders_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        selected_scrollbar = tk.Scrollbar(selected_frame, orient=tk.VERTICAL, command=self.selected_cylinders_listbox.yview)
        self.selected_cylinders_listbox.config(yscrollcommand=selected_scrollbar.set)
        selected_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.selected_cylinders_listbox.bind('<Double-1>', self.remove_selected_cylinder)

        # Configure grid weights for dispatch frame
        dispatch_frame.grid_columnconfigure(1, weight=1)

        ttk.Button(dispatch_frame, text="Dispatch Selected", command=self.dispatch_cylinders).grid(row=8, column=0, columnspan=2, pady=10)

        # Return section
        return_frame = ttk.LabelFrame(scrollable_frame, text="Return Cylinders", padding=5)
        return_frame.pack(fill=tk.X, padx=5, pady=5)

        tk.Label(return_frame, text="DC Number:").grid(row=0, column=0, padx=5, pady=3, sticky="w")
        self.dc_var = tk.StringVar()
        self.dc_combo = ttk.Combobox(return_frame, textvariable=self.dc_var,
                                     values=[], state="readonly", width=28)
        self.dc_combo.grid(row=0, column=1, padx=5, pady=3, sticky="ew")
        self.dc_combo.bind('<<ComboboxSelected>>', self.on_dc_select)

        tk.Label(return_frame, text="Dispatched Cylinders:").grid(row=1, column=0, padx=5, pady=3, sticky="nw")
        # Frame for listbox and scrollbar
        return_cylinder_frame = tk.Frame(return_frame)
        return_cylinder_frame.grid(row=1, column=1, padx=5, pady=3, sticky="ew")
        self.return_cylinder_listbox = tk.Listbox(return_cylinder_frame, selectmode=tk.MULTIPLE, height=4, width=32)
        self.return_cylinder_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        return_scrollbar = tk.Scrollbar(return_cylinder_frame, orient=tk.VERTICAL, command=self.return_cylinder_listbox.yview)
        self.return_cylinder_listbox.config(yscrollcommand=return_scrollbar.set)
        return_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        tk.Label(return_frame, text="Return Date:").grid(row=2, column=0, padx=5, pady=3, sticky="w")
        self.return_date_entry = tk.Entry(return_frame, width=30)
        self.return_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.return_date_entry.grid(row=2, column=1, padx=5, pady=3, sticky="ew")

        # Configure grid weights for return frame
        return_frame.grid_columnconfigure(1, weight=1)

        ttk.Button(return_frame, text="Return Selected", command=self.return_cylinders).grid(row=3, column=0, columnspan=2, pady=10)

        # Right panel - History view
        right_panel = ttk.LabelFrame(operations_tab, text="Dispatch History", padding=5)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(2, 5), pady=5)

        # Filter controls
        filter_frame = ttk.Frame(right_panel)
        filter_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(filter_frame, text="Filter by Status:").pack(side=tk.LEFT, padx=(0, 5))

        self.filter_var = tk.StringVar(value="dispatched")
        filter_combo = ttk.Combobox(filter_frame, textvariable=self.filter_var,
                                    values=["All", "dispatched", "returned", "refill", "maintenance"],
                                    state="readonly", width=15)
        filter_combo.pack(side=tk.LEFT, padx=5)
        filter_combo.bind('<<ComboboxSelected>>', self.on_filter_change)

        ttk.Label(filter_frame, text="Filter by Company:").pack(side=tk.LEFT, padx=(5, 5))

        self.company_filter_var = tk.StringVar(value="All")
        self.company_filter_combo = ttk.Combobox(filter_frame, textvariable=self.company_filter_var,
                                                 values=["All"] + [f"{c[0]} - {c[1]}" for c in self.customers],
                                                 state="readonly", width=22)
        self.company_filter_combo.pack(side=tk.LEFT, padx=5)
        self.company_filter_combo.bind('<<ComboboxSelected>>', self.on_filter_change)

        ttk.Label(filter_frame, text="Filter by DC:").pack(side=tk.LEFT, padx=(5, 5))

        self.dc_filter_var = tk.StringVar(value="All")
        self.dc_filter_combo = ttk.Combobox(filter_frame, textvariable=self.dc_filter_var,
                                            values=["All"], state="readonly", width=15)
        self.dc_filter_combo.pack(side=tk.LEFT, padx=5)
        self.dc_filter_combo.bind('<<ComboboxSelected>>', self.on_filter_change)

        # Button frame for right side buttons
        btn_frame = ttk.Frame(filter_frame)
        btn_frame.pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Return Selected", command=self.return_selected_cylinders).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Generate Bill", command=self.generate_bill).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.RIGHT, padx=2)
        ttk.Button(btn_frame, text="Refresh", command=self.load_dispatches).pack(side=tk.RIGHT, padx=2)

        # Treeview container with scrollbars
        tree_container = ttk.Frame(right_panel)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Treeview for dispatches
        columns = ('Select', 'ID', 'DC Number', 'Customer', 'Cylinder ID', 'Cylinder Type', 'Grade', 'Dispatch Date', 'Return Date', 'Status', 'Delete')
        self.tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=15)

        # Style the treeview
        style = ttk.Style()
        style.configure("Treeview", font=('Arial', 9), rowheight=25)
        style.configure("Treeview.Heading", font=('Arial', 9, 'bold'))
        
        # Configure colors for selected rows
        style.map('Treeview', background=[('selected', '#bbdefb')])
        
        # Configure tags for selection column
        self.tree.tag_configure('selected', foreground='#1976D2', font=('Arial', 12, 'bold'))
        self.tree.tag_configure('unselected', foreground="#000000", font=('Arial', 12))
        
        # Configure tags for status-based coloring in dispatch history
        self.tree.tag_configure('dispatched', foreground='#EF6C00')  # Orange
        self.tree.tag_configure('returned', foreground='#1565C0')    # Blue
        self.tree.tag_configure('refill', foreground='#6A1B9A')      # Purple
        self.tree.tag_configure('maintenance', foreground='#546E7A') # Gray

        for col in columns:
            self.tree.heading(col, text=col)
            if col == 'Select':
                self.tree.column(col, width=50, anchor='center')
            elif col == 'ID':
                self.tree.column(col, width=40, anchor='center')
            elif col == 'DC Number':
                self.tree.column(col, width=90)
            elif col == 'Cylinder Type':
                self.tree.column(col, width=90)
            elif col == 'Grade':
                self.tree.column(col, width=70)
            elif col == 'Delete':
                self.tree.column(col, width=50, anchor='center')
            else:
                self.tree.column(col, width=95)

        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Grid treeview and scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        # Configure grid weights
        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)

        # Bind double-click for delete
        self.tree.bind('<Double-1>', self.on_tree_double_click)
        # Bind click for select
        self.tree.bind('<Button-1>', self.on_tree_click)

        # Tab 2: Available Cylinders with Dispatch History
        available_tab = ttk.Frame(self.notebook)
        self.notebook.add(available_tab, text="Available Cylinders")
        
        # Container for available cylinders tab
        available_container = ttk.LabelFrame(available_tab, text="Available Cylinders with Dispatch History", padding=5)
        available_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Filter controls for available cylinders
        filter_cyl_frame = ttk.Frame(available_container)
        filter_cyl_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(filter_cyl_frame, text="Filter by Status:").pack(side=tk.LEFT, padx=(0, 5))
        self.available_filter_var = tk.StringVar(value="available")
        self.available_filter_combo = ttk.Combobox(filter_cyl_frame, textvariable=self.available_filter_var,
                                                     values=["All", "available", "dispatched", "returned"],
                                                     state="readonly", width=15)
        self.available_filter_combo.pack(side=tk.LEFT, padx=5)
        self.available_filter_combo.bind('<<ComboboxSelected>>', self.load_available_cylinders_history)
        
        ttk.Button(filter_cyl_frame, text="Refresh", command=self.load_available_cylinders_history).pack(side=tk.RIGHT, padx=5)
        
        # Treeview container with scrollbars
        cyl_tree_container = ttk.Frame(available_container)
        cyl_tree_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Treeview for available cylinders with history
        self.cyl_history_tree = ttk.Treeview(cyl_tree_container, columns=('Select', 'Cylinder ID', 'Type', 'Status', 'Current Location', 'Last DC', 'Last Customer', 'Last Dispatch Date', 'Last Return Date', 'Last Grade'), show='headings', height=15)
        
        # Style the treeview
        style = ttk.Style()
        style.configure("Treeview", font=('Arial', 9), rowheight=25)
        style.configure("Treeview.Heading", font=('Arial', 9, 'bold'))
        
        # Configure colors for selected rows
        style.map('Treeview', background=[('selected', '#bbdefb')])
        
        # Configure tags for selection column
        self.cyl_history_tree.tag_configure('selected', foreground='#1976D2', font=('Arial', 12, 'bold'))
        self.cyl_history_tree.tag_configure('unselected', foreground="#000000", font=('Arial', 12))
        
        # Configure tags for status-based coloring in available cylinders history
        self.cyl_history_tree.tag_configure('available', foreground='#2E7D32')  # Green
        self.cyl_history_tree.tag_configure('dispatched', foreground='#EF6C00')  # Orange
        self.cyl_history_tree.tag_configure('returned', foreground='#1565C0')    # Blue
        self.cyl_history_tree.tag_configure('refill', foreground='#6A1B9A')      # Purple
        self.cyl_history_tree.tag_configure('maintenance', foreground='#546E7A') # Gray
        
        for col in self.cyl_history_tree['columns']:
            self.cyl_history_tree.heading(col, text=col)
            if col == 'Select':
                self.cyl_history_tree.column(col, width=50, anchor='center')
            elif col == 'Cylinder ID':
                self.cyl_history_tree.column(col, width=100)
            elif col == 'Type':
                self.cyl_history_tree.column(col, width=80)
            elif col == 'Status':
                self.cyl_history_tree.column(col, width=80)
            elif col == 'Current Location':
                self.cyl_history_tree.column(col, width=100)
            elif col == 'Last Grade':
                self.cyl_history_tree.column(col, width=80)
            else:
                self.cyl_history_tree.column(col, width=110)
        
        # Scrollbars for available cylinders treeview
        cyl_v_scrollbar = ttk.Scrollbar(cyl_tree_container, orient=tk.VERTICAL, command=self.cyl_history_tree.yview)
        cyl_h_scrollbar = ttk.Scrollbar(cyl_tree_container, orient=tk.HORIZONTAL, command=self.cyl_history_tree.xview)
        self.cyl_history_tree.configure(yscrollcommand=cyl_v_scrollbar.set, xscrollcommand=cyl_h_scrollbar.set)
        
        # Grid treeview and scrollbars
        self.cyl_history_tree.grid(row=0, column=0, sticky="nsew")
        cyl_v_scrollbar.grid(row=0, column=1, sticky="ns")
        cyl_h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        cyl_tree_container.grid_columnconfigure(0, weight=1)
        cyl_tree_container.grid_rowconfigure(0, weight=1)
        
        # Bind click for select
        self.cyl_history_tree.bind('<Button-1>', self.on_cyl_history_click)
        
        # Return button for available cylinders tab
        return_btn_frame = ttk.Frame(available_container)
        return_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(return_btn_frame, text="Return Selected Cylinders", command=self.return_from_cyl_history).pack(side=tk.RIGHT, padx=5)
        
        # Configure tags for status-based coloring in available cylinders history
        self.cyl_history_tree.tag_configure('available', foreground='#2E7D32')  # Green
        self.cyl_history_tree.tag_configure('dispatched', foreground='#EF6C00')  # Orange
        self.cyl_history_tree.tag_configure('returned', foreground='#1565C0')    # Blue
        self.cyl_history_tree.tag_configure('refill', foreground='#6A1B9A')      # Purple
        self.cyl_history_tree.tag_configure('maintenance', foreground='#546E7A') # Gray

    def load_available_cylinders_history(self, event=None):
        """Load available cylinders with their dispatch history for reference during dispatch."""
        # Check if the filter variable exists (may not exist if called before widgets are created)
        if not hasattr(self, 'available_filter_var'):
            return
        
        filter_status = self.available_filter_var.get()
        
        # Clear current items
        for item in self.cyl_history_tree.get_children():
            self.cyl_history_tree.delete(item)
        
        conn = get_connection()
        cursor = conn.cursor()
        
        # Get all cylinders with their current status
        if filter_status == "All":
            cursor.execute("SELECT id, cylinder_id, cylinder_type, status, location FROM cylinders ORDER BY cylinder_id")
        else:
            cursor.execute("SELECT id, cylinder_id, cylinder_type, status, location FROM cylinders WHERE status = ? ORDER BY cylinder_id", (filter_status,))
        
        cylinders = cursor.fetchall()
        
        for cyl_id, cylinder_id_text, cyl_type, status, location in cylinders:
            # Get last dispatch history for this cylinder
            cursor.execute('''
                SELECT d.dc_number, c.name, d.dispatch_date, d.return_date, d.grade
                FROM dispatches d
                JOIN customers c ON d.customer_id = c.id
                WHERE d.cylinder_id = ?
                ORDER BY d.dispatch_date DESC
                LIMIT 1
            ''', (cyl_id,))
            history = cursor.fetchone()
            
            if history:
                last_dc, last_customer, last_dispatch_date, last_return_date, last_grade = history
            else:
                last_dc = "N/A"
                last_customer = "N/A"
                last_dispatch_date = "N/A"
                last_return_date = "N/A"
                last_grade = "N/A"
            
            current_location = location if location else "Warehouse"
            
            # Determine if this cylinder is selected
            cyl_id_str = str(cyl_id)
            is_selected = cyl_id_str in self.cyl_history_selected
            
            # Store tags with status, cylinder ID, and selection state
            status_tag = status
            selection_tag = 'selected' if is_selected else 'unselected'
            self.cyl_history_tree.insert('', tk.END, values=(
                '✓' if is_selected else '',  # Select column
                cylinder_id_text,
                cyl_type,
                status,
                current_location,
                last_dc,
                last_customer,
                last_dispatch_date,
                last_return_date,
                last_grade or 'N/A'
            ), tags=(status_tag, cyl_id_str, selection_tag))  # Store status, cyl_id, and selection as tags
        
        conn.close()

    def load_customers(self):
        """Load customers for dispatch combo."""
        self.customers = get_all_customers()
        # Update combo boxes if they exist
        if hasattr(self, 'customer_combo'):
            self.customer_combo['values'] = [f"{c[0]} - {c[1]}" for c in self.customers]
        if hasattr(self, 'company_filter_combo'):
            self.company_filter_combo['values'] = ["All"] + [f"{c[0]} - {c[1]}" for c in self.customers]

    def load_available_cylinders(self):
        """Load available cylinders for dispatch listbox."""
        self.available_cylinders = get_cylinders_by_status('available')
        # Update the listbox display
        if hasattr(self, 'cylinder_listbox'):
            self.cylinder_listbox.delete(0, tk.END)
            for cylinder in self.available_cylinders:
                self.cylinder_listbox.insert(tk.END, f"{cylinder[0]} - {cylinder[1]} ({cylinder[2]})")
        
        # Remove any dispatched cylinders from the selected cylinders listbox
        if hasattr(self, 'selected_cylinders_listbox'):
            # Get currently selected cylinder IDs from the listbox
            selected_to_remove = []
            for i in range(self.selected_cylinders_listbox.size()):
                cylinder_text = self.selected_cylinders_listbox.get(i)
                cylinder_id = int(cylinder_text.split(' - ')[0])
                # Check if this cylinder is still available
                if not any(c[0] == cylinder_id for c in self.available_cylinders):
                    selected_to_remove.append(i)
            
            # Remove in reverse order to preserve indices
            for index in reversed(selected_to_remove):
                self.selected_cylinders_listbox.delete(index)
        
        # Also refresh the available cylinders history (only if widgets exist)
        if hasattr(self, 'cyl_history_tree'):
            self.load_available_cylinders_history()

    def load_dispatches(self):
        """Load dispatches from database."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.dispatches = get_all_dispatches()
        for dispatch_row in self.dispatches:
            dispatch = Dispatch.from_db_row(dispatch_row)
            delete_text = 'Delete' if dispatch.status == 'returned' else ''
            is_selected = dispatch.id in self.selected_items
            select_text = '✓' if is_selected else ''
            tags = (dispatch.status, 'selected' if is_selected else 'unselected')
            values = (select_text, dispatch.id, dispatch.dc_number, dispatch.customer_name, dispatch.cylinder_id_text,
                      dispatch.cylinder_type, dispatch.grade or '', dispatch.dispatch_date, dispatch.return_date, dispatch.status, delete_text)
            self.tree.insert('', tk.END, values=values, tags=tags)

        # Update DC combo with unique DC numbers that have dispatched cylinders
        dc_numbers = list(set(d[1] for d in self.dispatches if d[8] == 'dispatched'))
        dc_numbers.sort(reverse=True)
        self.dc_combo['values'] = dc_numbers

        # Update DC filter combo
        all_dc_numbers = list(set(d[1] for d in self.dispatches))
        all_dc_numbers.sort(reverse=True)
        self.dc_filter_combo['values'] = ["All"] + all_dc_numbers

        # Apply current filters
        self.on_filter_change()
        
        # Also refresh the available cylinders history (only if widgets exist)
        if hasattr(self, 'cyl_history_tree'):
            self.load_available_cylinders_history()

    def generate_bill(self):
        """Generate a bill for the selected DC or company."""
        company_selection = self.company_filter_var.get()
        dc_selection = self.dc_filter_var.get()

        if dc_selection != "All":
            # Generate bill for specific DC
            bill_data = self.get_bill_data_for_dc(dc_selection)
            if not bill_data:
                messagebox.showinfo("No Data", f"No dispatches found for DC {dc_selection}.")
                return
            customer_name = bill_data['customer_name']
            bill_title = f"Bill for DC {dc_selection} - {customer_name}"
        elif company_selection != "All":
            # Generate bill for specific company
            customer_id = int(company_selection.split(' - ')[0])
            customer_name = company_selection.split(' - ')[1]
            bill_data = self.get_bill_data_for_company(customer_id)
            if not bill_data:
                messagebox.showinfo("No Data", f"No dispatches found for {customer_name}.")
                return
            bill_title = f"Bill for {customer_name}"
        else:
            messagebox.showerror("Error", "Please select a specific DC or company to generate bill.")
            return

        # Generate PDF bill
        self.create_pdf_bill(bill_title, bill_data)

    def get_bill_data_for_dc(self, dc_number):
        """Get bill data for a specific DC number."""
        conn = get_connection()
        cursor = conn.cursor()
        # Get customer info from dispatches
        cursor.execute('''
            SELECT DISTINCT c.id, c.name, c.contact_info, c.address
            FROM dispatches d
            JOIN customers c ON d.customer_id = c.id
            WHERE d.dc_number = ?
        ''', (dc_number,))
        customer_row = cursor.fetchone()
        if not customer_row:
            conn.close()
            return None
        customer_id, customer_name, contact_info, address = customer_row

        # Get dispatches
        cursor.execute('''
            SELECT d.dc_number, d.dispatch_date, d.return_date, d.status, cy.cylinder_id, cy.cylinder_type, d.grade, d.dispatch_notes, d.return_notes
            FROM dispatches d
            JOIN cylinders cy ON d.cylinder_id = cy.id
            WHERE d.dc_number = ?
            ORDER BY d.dispatch_date DESC
        ''', (dc_number,))
        dispatches = cursor.fetchall()
        conn.close()

        return {
            'customer_id': customer_id,
            'customer_name': customer_name,
            'contact_info': contact_info,
            'address': address,
            'dispatches': dispatches
        }

    def create_pdf_bill(self, bill_title, bill_data):
        """Create a professional PDF bill."""
        if SimpleDocTemplate is None:
            messagebox.showerror("Error", "reportlab is not installed. Cannot generate PDF.")
            return

        # Default filename
        customer_name_clean = bill_data['customer_name'].replace(' ', '_')
        default_name = f"{customer_name_clean}_Bill_{datetime.now().strftime('%Y-%m-%d')}.pdf"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            title="Save Bill as PDF",
            initialfile=default_name
        )
        if not file_path:
            return

        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = styles['Heading1']
            title_style.alignment = 1  # Center
            story.append(Paragraph(bill_title, title_style))
            story.append(Spacer(1, 12))

            # Company Header
            company_info = f"<b>Customer:</b> {bill_data['customer_name']}<br/>"
            if bill_data['contact_info']:
                company_info += f"<b>Contact:</b> {bill_data['contact_info']}<br/>"
            if bill_data['address']:
                company_info += f"<b>Address:</b> {bill_data['address']}<br/>"
            story.append(Paragraph(company_info, styles['Normal']))
            story.append(Spacer(1, 12))

            # Generated on
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 12))

            # Table data
            data = [['DC Number', 'Cylinder ID', 'Type', 'Grade', 'Dispatch Date', 'Return Date', 'Status']]
            total_cylinders = 0
            dispatched_count = 0
            returned_count = 0

            for dc, disp_date, ret_date, status, cyl_id, cyl_type, grade, disp_notes, ret_notes in bill_data['dispatches']:
                data.append([dc, cyl_id, cyl_type, grade or '', disp_date, ret_date or '', status])
                total_cylinders += 1
                if status == 'dispatched':
                    dispatched_count += 1
                elif status == 'returned':
                    returned_count += 1

            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))

            # Summary
            summary = f"""
            <b>Summary:</b><br/>
            Total Cylinders: {total_cylinders}<br/>
            Currently Dispatched: {dispatched_count}<br/>
            Returned: {returned_count}
            """
            story.append(Paragraph(summary, styles['Normal']))

            # Build PDF
            doc.build(story)

            messagebox.showinfo("Success", f"Bill saved to {file_path}")

            # Ask to print
            if messagebox.askyesno("Print Bill", "Do you want to print the bill?"):
                self.print_pdf(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {e}")

    def print_pdf(self, file_path):
        """Print the PDF file."""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_path, "print")
            else:
                # For other OS, use system print command
                os.system(f"lpr {file_path}")
        except Exception as e:
            messagebox.showerror("Print Error", f"Failed to print: {e}")

    def get_bill_data_for_company(self, customer_id):
        """Get bill data for a specific company."""
        conn = get_connection()
        cursor = conn.cursor()
        # Get customer info
        cursor.execute('SELECT name, contact_info, address FROM customers WHERE id = ?', (customer_id,))
        customer_row = cursor.fetchone()
        if not customer_row:
            conn.close()
            return None
        customer_name, contact_info, address = customer_row

        # Get dispatches
        cursor.execute('''
            SELECT d.dc_number, d.dispatch_date, d.return_date, d.status, cy.cylinder_id, cy.cylinder_type, d.grade, d.dispatch_notes, d.return_notes
            FROM dispatches d
            JOIN cylinders cy ON d.cylinder_id = cy.id
            WHERE d.customer_id = ?
            ORDER BY d.dc_number DESC, d.dispatch_date DESC
        ''', (customer_id,))
        dispatches = cursor.fetchall()
        conn.close()

        return {
            'customer_id': customer_id,
            'customer_name': customer_name,
            'contact_info': contact_info,
            'address': address,
            'dispatches': dispatches
        }

    def export_to_excel(self):
        """Export the current dispatch history to Excel."""
        if Workbook is None:
            messagebox.showerror("Error", "openpyxl is not installed. Cannot export to Excel.")
            return

        # Default filename based on selected company and date
        company_selection = self.company_filter_var.get()
        if company_selection == "All":
            default_name = f"Dispatch_History_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        else:
            company_name = company_selection.split(' - ')[1].replace(' ', '_')
            default_name = f"{company_name}_Dispatch_History_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

        # Get file path from user
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Dispatch History",
            initialfile=default_name
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Dispatch History"

            # Headers
            headers = ['Select', 'ID', 'DC Number', 'Customer', 'Cylinder ID', 'Cylinder Type', 'Grade', 'Dispatch Date', 'Return Date', 'Status', 'Delete']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Data from treeview
            for row_num, item in enumerate(self.tree.get_children(), 2):
                values = self.tree.item(item, 'values')
                for col_num, value in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

            wb.save(file_path)
            messagebox.showinfo("Success", f"Data exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")

    def on_filter_change(self, event=None):
        """Handle filter change."""
        filter_status = self.filter_var.get()
        filter_company = self.company_filter_var.get()
        filter_dc = self.dc_filter_var.get()
        # Clear current items
        for item in self.tree.get_children():
            self.tree.delete(item)

        filtered_dispatches = self.dispatches

        # Filter by status
        if filter_status != "All":
            filtered_dispatches = [d for d in filtered_dispatches if d[8] == filter_status]

        # Filter by company
        if filter_company != "All":
            company_id = int(filter_company.split(' - ')[0])
            filtered_dispatches = [d for d in filtered_dispatches if d[2] == company_id]

        # Filter by DC
        if filter_dc != "All":
            filtered_dispatches = [d for d in filtered_dispatches if d[1] == filter_dc]

        for dispatch_row in filtered_dispatches:
            dispatch = Dispatch.from_db_row(dispatch_row)
            delete_text = 'Delete' if dispatch.status == 'returned' else ''
            is_selected = dispatch.id in self.selected_items
            select_text = '✓' if is_selected else ''
            tags = (dispatch.status, 'selected' if is_selected else 'unselected')
            values = (select_text, dispatch.id, dispatch.dc_number, dispatch.customer_name, dispatch.cylinder_id_text,
                      dispatch.cylinder_type, dispatch.grade or '', dispatch.dispatch_date, dispatch.return_date, dispatch.status, delete_text)
            self.tree.insert('', tk.END, values=values, tags=tags)

    def on_dc_select(self, event=None):
        """Handle DC number selection for return."""
        dc_number = self.dc_var.get()
        if dc_number:
            # Load cylinders for this DC number
            self.return_cylinder_listbox.delete(0, tk.END)
            dispatched_cylinders = get_dispatched_cylinders_by_dc(dc_number)
            for cylinder in dispatched_cylinders:
                # Get cylinder type from database
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT cylinder_type FROM cylinders WHERE id = ?", (cylinder[0],))
                cylinder_type = cursor.fetchone()[0]
                conn.close()
                self.return_cylinder_listbox.insert(tk.END, f"{cylinder[0]} - {cylinder[1]} ({cylinder_type})")

    def resolve_cylinder_id(self, input_id):
        """Resolve input to cylinder database ID. Accepts ID or cylinder_id_text."""
        conn = get_connection()
        cursor = conn.cursor()
        try:
            # Try as integer ID
            try:
                cylinder_id = int(input_id)
                cursor.execute("SELECT id FROM cylinders WHERE id = ?", (cylinder_id,))
                result = cursor.fetchone()
                if result:
                    return cylinder_id
            except ValueError:
                pass

            # Try as cylinder_id_text
            cursor.execute("SELECT id FROM cylinders WHERE cylinder_id = ?", (input_id,))
            result = cursor.fetchone()
            if result:
                return result[0]

            return None
        finally:
            conn.close()

    def on_available_cylinder_select(self):
        """Handle cylinder selection from available listbox - add to selected list."""
        selected_indices = self.cylinder_listbox.curselection()
        
        # Get all currently selected cylinder IDs
        selected_cylinder_ids = set()
        for i in range(self.selected_cylinders_listbox.size()):
            cylinder_text = self.selected_cylinders_listbox.get(i)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            selected_cylinder_ids.add(cylinder_id)
        
        # Add newly selected cylinders
        for index in selected_indices:
            cylinder_text = self.cylinder_listbox.get(index)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            selected_cylinder_ids.add(cylinder_id)
        
        # Rebuild selected listbox with all selected cylinders
        self.selected_cylinders_listbox.delete(0, tk.END)
        if selected_cylinder_ids:
            conn = get_connection()
            cursor = conn.cursor()
            for cylinder_id in sorted(selected_cylinder_ids):
                cursor.execute("SELECT cylinder_id, cylinder_type FROM cylinders WHERE id = ? AND status = 'available'", (cylinder_id,))
                result = cursor.fetchone()
                if result:
                    self.selected_cylinders_listbox.insert(tk.END, f"{cylinder_id} - {result[0]} ({result[1]})")
            conn.close()

    def on_available_cylinder_deselect(self):
        """Handle cylinder deselection from available listbox - remove from selected list."""
        # Get currently selected indices in available listbox
        selected_indices = set(self.cylinder_listbox.curselection())
        
        # Get all cylinder IDs that are still selected in available listbox
        still_selected_ids = set()
        for index in selected_indices:
            cylinder_text = self.cylinder_listbox.get(index)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            still_selected_ids.add(cylinder_id)
        
        # Remove deselected cylinders from selected listbox
        # Go in reverse order to preserve indices while deleting
        for i in range(self.selected_cylinders_listbox.size() - 1, -1, -1):
            cylinder_text = self.selected_cylinders_listbox.get(i)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            if cylinder_id not in still_selected_ids:
                self.selected_cylinders_listbox.delete(i)

    def update_selected_cylinders(self):
        """Update the selected cylinders listbox with current selections."""
        #self.selected_cylinders_listbox.delete(0, tk.END)

        selected_indices = self.cylinder_listbox.curselection()
        manual_cylinders = self.manual_cylinder_entry.get().strip()

        cylinder_ids = set()
        errors = []

        # From listbox
        for index in selected_indices:
            cylinder_text = self.cylinder_listbox.get(index)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            cylinder_ids.add(cylinder_id)

        # From manual input
        if manual_cylinders:
            manual_ids = [id.strip() for id in manual_cylinders.split(',') if id.strip()]
            for manual_id in manual_ids:
                resolved_id = self.resolve_cylinder_id(manual_id)
                if resolved_id is None:
                    errors.append(f"Invalid cylinder: {manual_id}")
                else:
                    cylinder_ids.add(resolved_id)

        # Check all cylinder_ids for availability and existence
        valid_cylinder_ids = set()
        conn = get_connection()
        cursor = conn.cursor()
        for cylinder_id in cylinder_ids:
            cursor.execute("SELECT status, cylinder_id FROM cylinders WHERE id = ?", (cylinder_id,))
            result = cursor.fetchone()
            if result:
                status, cyl_id = result
                if status == 'available':
                    valid_cylinder_ids.add(cylinder_id)
                else:
                    errors.append(f"Cylinder {cyl_id} (ID: {cylinder_id}) is not available (status: {status})")
            else:
                errors.append(f"Cylinder ID {cylinder_id} not found")
        conn.close()

        # Display errors if any
        if errors:
            error_msg = "Errors in cylinder selection:\n" + "\n".join(errors)
            messagebox.showerror("Cylinder Selection Errors", error_msg)
        else:
            # Clear manual input field if added successfully
            if manual_cylinders:
                self.manual_cylinder_entry.delete(0, tk.END)

        # Display valid cylinders, avoiding duplicates
        if valid_cylinder_ids:
            # Get currently displayed cylinder IDs to avoid duplicates
            existing_cylinder_ids = set()
            for i in range(self.selected_cylinders_listbox.size()):
                cylinder_text = self.selected_cylinders_listbox.get(i)
                cylinder_id = int(cylinder_text.split(' - ')[0])
                existing_cylinder_ids.add(cylinder_id)
            
            conn = get_connection()
            cursor = conn.cursor()
            for cylinder_id in sorted(valid_cylinder_ids):
                if cylinder_id not in existing_cylinder_ids:
                    cursor.execute("SELECT cylinder_id, cylinder_type FROM cylinders WHERE id = ?", (cylinder_id,))
                    result = cursor.fetchone()
                    if result:
                        self.selected_cylinders_listbox.insert(tk.END, f"{cylinder_id} - {result[0]} ({result[1]})")
            conn.close()

    def remove_selected_cylinder(self, event):
        """Remove a cylinder from the selected list on double-click."""
        selection = self.selected_cylinders_listbox.curselection()
        if selection:
            index = selection[0]
            cylinder_text = self.selected_cylinders_listbox.get(index)
            cylinder_id = int(cylinder_text.split(' - ')[0])
            cylinder_display = cylinder_text.split(' - ')[1].split(' (')[0]  # Get the cylinder_id_text
            if messagebox.askyesno("Confirm Removal", f"Do you want to remove cylinder {cylinder_display} from the selected list?"):
                self.selected_cylinders_listbox.delete(index)

    def dispatch_cylinders(self):
        """Dispatch selected cylinders to a customer."""
        customer_selection = self.customer_var.get()
        dispatch_date = self.dispatch_date_entry.get().strip()
        grade = self.grade_entry.get().strip()
        dispatch_notes = ""

        if not customer_selection:
            messagebox.showerror("Error", "Please select a customer.")
            return

        if not grade:
            messagebox.showerror("Error", "Grade is mandatory. Please enter the grade.")
            return

        if self.selected_cylinders_listbox.size() == 0:
            messagebox.showerror("Error", "Please select cylinders to dispatch.")
            return

        if not dispatch_date:
            messagebox.showerror("Error", "Please enter dispatch date.")
            return

        try:
            # Extract customer ID
            customer_id = int(customer_selection.split(' - ')[0])

            # Extract cylinder IDs from selected cylinders listbox
            cylinder_ids = []
            for i in range(self.selected_cylinders_listbox.size()):
                cylinder_text = self.selected_cylinders_listbox.get(i)
                cylinder_id = int(cylinder_text.split(' - ')[0])
                cylinder_ids.append(cylinder_id)

            dc_number = self.dc_number_var.get().strip()
            if not dc_number:
                dc_number = None  # Will generate new one
            dc_number = dispatch_cylinders(customer_id, cylinder_ids, dispatch_date, dispatch_notes, dc_number, grade)
            self.load_dispatches()
            self.load_available_cylinders()

            self.customer_var.set('')
            # Do not clear DC number
            self.manual_cylinder_entry.delete(0, tk.END)
            self.grade_entry.delete(0, tk.END)
            self.selected_cylinders_listbox.delete(0, tk.END)  # Clear selected cylinders after successful dispatch
            messagebox.showinfo("Success", f"Cylinders dispatched successfully under DC {dc_number}.")
        except ValueError as e:
            messagebox.showerror("Validation Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to dispatch cylinders: {e}")

    def on_tree_click(self, event):
        """Handle click on treeview for select toggle."""
        region = self.tree.identify_region(event.x, event.y)
        if region == 'cell':
            item = self.tree.identify_row(event.y)
            if item:
                values = self.tree.item(item, 'values')
                dispatch_id = values[1]
                if dispatch_id in self.selected_items:
                    self.selected_items.remove(dispatch_id)
                    self.tree.item(item, tags=(self.tree.item(item, 'tags')[0], 'unselected'))
                else:
                    self.selected_items.add(dispatch_id)
                    self.tree.item(item, tags=(self.tree.item(item, 'tags')[0], 'selected'))
                # Update display
                select_text = '✓' if dispatch_id in self.selected_items else ''
                self.tree.set(item, 'Select', select_text)

    def on_tree_double_click(self, event):
        """Handle double-click on treeview for delete action."""
        item = self.tree.selection()
        if not item:
            return
        item = item[0]
        values = self.tree.item(item, 'values')
        if len(values) > 10 and values[10] == 'Delete':
            dispatch_id = values[1]
            # Confirm delete
            if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete dispatch record ID {dispatch_id}?"):
                try:
                    dispatch = Dispatch(id=dispatch_id)
                    dispatch.delete()
                    self.load_dispatches()
                    messagebox.showinfo("Success", "Dispatch record deleted successfully.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete dispatch record: {e}")

    def on_cyl_history_click(self, event):
        """Handle click on available cylinders history treeview for select toggle."""
        region = self.cyl_history_tree.identify_region(event.x, event.y)
        if region == 'cell':
            item = self.cyl_history_tree.identify_row(event.y)
            if item:
                values = self.cyl_history_tree.item(item, 'values')
                tags = self.cyl_history_tree.item(item, 'tags')
                # tags[0] = status, tags[1] = cyl_id, tags[2] = selection state
                status_tag = tags[0]
                cyl_id = tags[1]  # cyl_id is at index 1
                if cyl_id in self.cyl_history_selected:
                    self.cyl_history_selected.remove(cyl_id)
                    self.cyl_history_tree.item(item, tags=(status_tag, cyl_id, 'unselected'))
                else:
                    self.cyl_history_selected.add(cyl_id)
                    self.cyl_history_tree.item(item, tags=(status_tag, cyl_id, 'selected'))
                # Update display
                select_text = '✓' if cyl_id in self.cyl_history_selected else ''
                self.cyl_history_tree.set(item, 'Select', select_text)

    def return_from_cyl_history(self):
        """Return cylinders selected in the Available Cylinders history tab."""
        if not self.cyl_history_selected:
            messagebox.showerror("Error", "Please select at least one cylinder to return.")
            return
        
        return_date = datetime.now().strftime("%Y-%m-%d")
        return_notes = ""
        
        try:
            # Group selected cylinders by DC number
            dc_groups = {}
            selected_cylinder_ids = []
            
            for cyl_id in self.cyl_history_selected:
                # Get the last dispatch for this cylinder
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT d.dc_number, d.id
                    FROM dispatches d
                    WHERE d.cylinder_id = ? AND d.status = 'dispatched'
                    ORDER BY d.dispatch_date DESC
                    LIMIT 1
                ''', (int(cyl_id),))
                dispatch = cursor.fetchone()
                conn.close()
                
                if dispatch:
                    dc_number, dispatch_id = dispatch
                    if dc_number not in dc_groups:
                        dc_groups[dc_number] = []
                    dc_groups[dc_number].append(int(cyl_id))
                    selected_cylinder_ids.append(int(cyl_id))
                else:
                    messagebox.showerror("Error", f"Cylinder {cyl_id} is not currently dispatched.")
                    return
            
            # Confirmation dialog
            cylinder_ids_str = ', '.join(str(cid) for cid in selected_cylinder_ids)
            if not messagebox.askyesno("Confirm Return", f"Confirm return of cylinders: {cylinder_ids_str}?"):
                return
            
            # Now return for each DC
            for dc, cyl_ids in dc_groups.items():
                return_cylinders(dc, cyl_ids, return_date, return_notes)
            
            # Clear selection
            self.cyl_history_selected.clear()
            self.load_dispatches()
            self.load_available_cylinders()
            self.load_available_cylinders_history()
            messagebox.showinfo("Success", "Selected cylinders returned successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to return cylinders: {e}")

    def return_cylinders(self):
        """Return selected cylinders."""
        dc_number = self.dc_var.get()
        selected_indices = self.return_cylinder_listbox.curselection()
        return_date = self.return_date_entry.get().strip()
        return_notes = ""

        if not dc_number:
            messagebox.showerror("Error", "Please select a DC number.")
            return

        if not selected_indices:
            messagebox.showerror("Error", "Please select at least one cylinder to return.")
            return

        if not return_date:
            messagebox.showerror("Error", "Please enter return date.")
            return

        try:
            # Extract cylinder IDs
            cylinder_ids = []
            for index in selected_indices:
                cylinder_text = self.return_cylinder_listbox.get(index)
                cylinder_id = int(cylinder_text.split(' - ')[0])
                cylinder_ids.append(cylinder_id)

            # Confirmation dialog
            cylinder_ids_str = ', '.join(str(cid) for cid in cylinder_ids)
            if not messagebox.askyesno("Confirm Return", f"Confirm return of cylinders: {cylinder_ids_str}?"):
                return

            return_cylinders(dc_number, cylinder_ids, return_date, return_notes)
            self.load_dispatches()
            self.load_available_cylinders()

            # Refresh return section
            self.dc_var.set('')
            self.return_cylinder_listbox.delete(0, tk.END)
            messagebox.showinfo("Success", "Cylinders returned successfully.")
        except ValueError as e:
            messagebox.showerror("Validation Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to return cylinders: {e}")

    def return_selected_cylinders(self):
        """Return cylinders selected in the treeview."""
        if not self.selected_items:
            messagebox.showerror("Error", "Please select at least one cylinder to return.")
            return

        return_date = datetime.now().strftime("%Y-%m-%d")  # Use current date
        return_notes = ""

        try:
            # Group selected items by DC number
            dc_groups = {}
            selected_cylinder_ids = []
            for item in self.tree.get_children():
                values = self.tree.item(item, 'values')
                dispatch_id = values[1]
                if dispatch_id in self.selected_items:
                    dc_number = values[2]
                    for d in self.dispatches:
                        if d[0] == dispatch_id and d[8] == 'dispatched':  # Only process dispatched cylinders
                            cylinder_id = d[3]  # d.cylinder_id
                            if dc_number not in dc_groups:
                                dc_groups[dc_number] = []
                            dc_groups[dc_number].append(cylinder_id)
                            selected_cylinder_ids.append(cylinder_id)
                            break

            # Confirmation dialog
            cylinder_ids_str = ', '.join(str(cid) for cid in selected_cylinder_ids)
            if not messagebox.askyesno("Confirm Return", f"Confirm return of cylinders: {cylinder_ids_str}?"):
                return

            # Now return for each DC
            for dc, cyl_ids in dc_groups.items():
                return_cylinders(dc, cyl_ids, return_date, return_notes)

            self.selected_items.clear()
            self.load_dispatches()
            self.load_available_cylinders()
            messagebox.showinfo("Success", "Selected cylinders returned successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to return cylinders: {e}")
